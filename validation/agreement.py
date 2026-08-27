"""Agreement between AI-suggested codes (column J) and the authors' final codes (column L)."""
import math
from openpyxl import load_workbook
wb = load_workbook('Provjera_preciznosti_upita.xlsx', data_only=True)
pairs = []
for sh, n in [('Uzorak_100', 100), ('Uzorak_50_rano', 50)]:
    ws = wb[sh]
    for r in range(5, 5 + n):
        pairs.append((str(ws.cell(row=r, column=10).value).strip().upper(), str(ws.cell(row=r, column=12).value).strip().upper()))
N = len(pairs); po = sum(a == b for a, b in pairs) / N
cats = 'ABC'; pe = sum((sum(a == c for a, _ in pairs) / N) * (sum(b == c for _, b in pairs) / N) for c in cats)
idx = {c: i for i, c in enumerate(cats)}
O = [[0] * 3 for _ in range(3)]
for a, b in pairs: O[idx[a]][idx[b]] += 1
ra = [sum(O[i]) for i in range(3)]; cb = [sum(O[i][j] for i in range(3)) for j in range(3)]
wo = sum(abs(i - j) * O[i][j] for i in range(3) for j in range(3)) / N
we = sum(abs(i - j) * ra[i] * cb[j] / N / N for i in range(3) for j in range(3))
print(f'N={N}  agreement={po:.3f}  kappa={(po-pe)/(1-pe):.3f}  weighted kappa={1-wo/we:.3f}')
print('confusion (rows = AI, cols = authors, order A B C):', O)
