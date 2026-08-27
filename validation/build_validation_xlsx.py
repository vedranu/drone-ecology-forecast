# -*- coding: utf-8 -*-
"""Build Provjera_preciznosti_upita.xlsx: two random samples with suggested codes, a final-code column
for the authors, and a summary sheet with live formulas (precision, Wilson CI, early vs recent test,
strict-query sensitivity)."""
import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from classify import S100, S50

FONT = 'Arial'
b = Font(name=FONT, size=10, bold=True); n = Font(name=FONT, size=10); small = Font(name=FONT, size=9)
blue = Font(name=FONT, size=10, color='0000FF'); hdr = PatternFill('solid', fgColor='D9E2F3'); yellow = PatternFill('solid', fgColor='FFFF00')
thin = Side(style='thin', color='BFBFBF')
D_TERMS = ['drone', 'drones', 'uav', 'uavs', 'uas', 'rpas', 'unmanned aerial', 'unmanned aircraft', 'remotely piloted aircraft']

def hit_terms(r):
    txt = (r['title'] + ' ' + r['abstract']).lower()
    return ', '.join(t for t in D_TERMS if re.search(r'\b' + re.escape(t) + r'\b', txt)) or '(nije u naslovu/sažetku)'

wb = Workbook()

# ---------------- Kriteriji ----------------
ws = wb.active; ws.title = 'Kriteriji'
lines = [
 ('Ručna provjera preciznosti upita D AND E (OpenAlex) — kriteriji kodiranja', True),
 ('', False),
 ('Svaki zapis dobiva jednu od tri oznake:', False),
 ('A — RELEVANTAN: bespilotna letjelica (UAV/UAS/dron) je senzorska platforma ILI predmet rada, A primjena je ekološka / okolišna: divlje životinje, staništa, šume i vegetacija u prirodnom okolišu, vode, zrak i emisije plinova, požari, obala i geomorfologija okoliša, onečišćenje, zaštićena područja, bioraznolikost.', False),
 ('B — GRANIČAN: UAV jest platforma, ali primjena je susjedna: precizna poljoprivreda (usjevi, prinos, korovi, fenotipizacija), urbanistička/geodetska izmjera, hazardi bez ekološkog fokusa, generička UAV tehnologija koja okoliš samo spominje.', False),
 ('C — NERELEVANTAN: nema zračne bespilotne letjelice („drone“ = trut, „UAS“ = drugi akronim, podvodni dronovi, književnost) ILI UAV bez ikakve okolišne primjene (vojna, dostava, sigurnost, softver).', False),
 ('', False),
 ('Postupak: (1) pročitajte naslov i sažetak; (2) ako sažetak nedostaje, otvorite poveznicu (OpenAlex ID) i pročitajte rad; (3) u stupac „KONAČNA OCJENA“ upišite A, B ili C — stupac je unaprijed popunjen prijedlogom (AI), koji je samo pomoć i može se promijeniti; (4) po potrebi napišite napomenu. List „Sažetak“ se sam preračunava.', False),
 ('', False),
 ('Uzorci: (a) 100 slučajnih zapisa iz cijelog razdoblja 2005.–2025. (OpenAlex sample=100, seed=20260826); (b) 50 slučajnih zapisa iz ranog razdoblja 2005.–2015. (seed=20260827) za provjeru stabilnosti preciznosti kroz vrijeme. Dohvat 27. 8. 2026.', False),
 ('', False),
 ('Za rad se izvještava: preciznost u strogom smislu P(A) i u širem smislu P(A ili B), s Wilsonovim 95 % intervalom; test razlike ranog i cijelog uzorka; te osjetljivost trenda na strogi upit (D AND E NOT poljoprivredni pojmovi), list „Strogi_upit“.', False),
 ('', False),
 ('Preporuka za drugu osobu (inter-rater): neka koautor neovisno kodira barem 30 zapisa u praznoj kopiji stupca; slaganje se izvještava kao postotak ili Cohenova kappa.', False),
]
for i, (t, bold) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=t); c.font = Font(name=FONT, size=11 if bold else 10, bold=bold); c.alignment = Alignment(wrap_text=True, vertical='top')
ws.column_dimensions['A'].width = 150

# ---------------- sample sheets ----------------
def sample_sheet(name, rows, codes, title):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=title).font = Font(name=FONT, size=11, bold=True)
    ws.cell(row=2, column=1, value='Žuti stupac = KONAČNA OCJENA (A/B/C), popunite ili potvrdite; prijedlog (AI) je samo pomoć.').font = Font(name=FONT, size=9, italic=True)
    head = ['#', 'Godina', 'Naslov', 'Časopis / izvor', 'OpenAlex tema', 'OpenAlex ID (poveznica)', 'DOI', 'UAV pojam', 'Sažetak (skraćen)', 'Prijedlog (AI)', 'Obrazloženje prijedloga', 'KONAČNA OCJENA', 'Napomena ocjenjivača']
    for c, h in enumerate(head, 1):
        cell = ws.cell(row=4, column=c, value=h); cell.font = b; cell.fill = hdr; cell.alignment = Alignment(wrap_text=True, vertical='center'); cell.border = Border(bottom=thin)
    for i, r in enumerate(rows, start=5):
        code, reason = codes[r['n']]
        vals = [r['n'], r['year'], r['title'] or '(bez naslova)', r['venue'], r['topic'], r['id'], r['doi'], hit_terms(r), (r['abstract'][:700] + ('…' if len(r['abstract']) > 700 else '')) or '(sažetak nije dostupan — otvoriti poveznicu)', code, reason, code, '']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v); cell.font = n; cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=i, column=6).hyperlink = r['id']; ws.cell(row=i, column=6).font = Font(name=FONT, size=10, color='0563C1', underline='single')
        ws.cell(row=i, column=12).fill = yellow; ws.cell(row=i, column=12).alignment = Alignment(horizontal='center', vertical='top')
        ws.row_dimensions[i].height = 75
    last = 4 + len(rows)
    dv = DataValidation(type='list', formula1='"A,B,C"', allow_blank=True); ws.add_data_validation(dv); dv.add(f'L5:L{last}')
    for c, w in enumerate([4, 7, 40, 22, 22, 26, 20, 14, 70, 9, 34, 11, 26], 1): ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'D5'; ws.auto_filter.ref = f'A4:M{last}'
    return last

rows100 = json.load(open('sample_rows.json')); rows50 = json.load(open('sample_rows_early.json'))
l100 = sample_sheet('Uzorak_100', rows100, S100, 'Uzorak A: 100 slučajnih zapisa, 2005.–2025. (seed 20260826)')
l50 = sample_sheet('Uzorak_50_rano', rows50, S50, 'Uzorak B: 50 slučajnih zapisa, 2005.–2015. (seed 20260827)')

# ---------------- summary with formulas ----------------
ws = wb.create_sheet('Sažetak', 1)
ws.cell(row=1, column=1, value='Sažetak provjere preciznosti (sve formule; mijenja se automatski kad promijenite KONAČNU OCJENU)').font = Font(name=FONT, size=11, bold=True)
def block(r0, sheet, last, label):
    rng = f"'{sheet}'!$L$5:$L${last}"
    ws.cell(row=r0, column=1, value=label).font = b
    labels = ['n (ocijenjeno)', 'A — relevantan', 'B — graničan', 'C — nerelevantan', 'P(A) stroga preciznost', 'P(A ili B) šira preciznost', 'Wilson 95 % donja, P(A)', 'Wilson 95 % gornja, P(A)', 'Wilson 95 % donja, P(A∪B)', 'Wilson 95 % gornja, P(A∪B)']
    for i, lab in enumerate(labels, 1): ws.cell(row=r0 + i, column=1, value=lab).font = n
    nA, nB, nC, nn = f'B{r0+2}', f'B{r0+3}', f'B{r0+4}', f'B{r0+1}'
    ws.cell(row=r0 + 1, column=2, value=f'=COUNTIF({rng},"A")+COUNTIF({rng},"B")+COUNTIF({rng},"C")')
    ws.cell(row=r0 + 2, column=2, value=f'=COUNTIF({rng},"A")'); ws.cell(row=r0 + 3, column=2, value=f'=COUNTIF({rng},"B")'); ws.cell(row=r0 + 4, column=2, value=f'=COUNTIF({rng},"C")')
    ws.cell(row=r0 + 5, column=2, value=f'=IF({nn}>0,{nA}/{nn},"")'); ws.cell(row=r0 + 6, column=2, value=f'=IF({nn}>0,({nA}+{nB})/{nn},"")')
    def wilson(p, N, lo):
        z = 1.96; sgn = '-' if lo else '+'
        return f'=IF({N}>0,(({p}+{z}^2/(2*{N})){sgn}{z}*SQRT({p}*(1-{p})/{N}+{z}^2/(4*{N}^2)))/(1+{z}^2/{N}),"")'
    ws.cell(row=r0 + 7, column=2, value=wilson(f'B{r0+5}', nn, True)); ws.cell(row=r0 + 8, column=2, value=wilson(f'B{r0+5}', nn, False))
    ws.cell(row=r0 + 9, column=2, value=wilson(f'B{r0+6}', nn, True)); ws.cell(row=r0 + 10, column=2, value=wilson(f'B{r0+6}', nn, False))
    for i in range(1, 11):
        c = ws.cell(row=r0 + i, column=2); c.font = n; c.number_format = '0' if i <= 4 else '0.0%'
    return r0 + 11
rA = 3; r = block(rA, 'Uzorak_100', l100, 'Uzorak A (2005.–2025., n = 100)')
rB = r + 1; r = block(rB, 'Uzorak_50_rano', l50, 'Uzorak B (2005.–2015., n = 50)')
nA_, aA_, pA_ = f'B{rA+1}', f'B{rA+2}', f'B{rA+5}'; nB_, aB_, pB_ = f'B{rB+1}', f'B{rB+2}', f'B{rB+5}'
# two-proportion z test A vs B for P(A)
ws.cell(row=r + 1, column=1, value='Razlika P(A): uzorak B (rano) − uzorak A (cijelo razdoblje)').font = b
ws.cell(row=r + 2, column=1, value='razlika udjela').font = n; ws.cell(row=r + 2, column=2, value=f'={pB_}-{pA_}').number_format = '0.0%'
ws.cell(row=r + 3, column=1, value='z (dvoproporcijski test, združeni p)').font = n
ws.cell(row=r + 3, column=2, value=f'=IF(AND({nA_}>0,{nB_}>0),({pB_}-{pA_})/SQRT((({aA_}+{aB_})/({nA_}+{nB_}))*(1-({aA_}+{aB_})/({nA_}+{nB_}))*(1/{nA_}+1/{nB_})),"")').number_format = '0.00'
ws.cell(row=r + 4, column=1, value='p (dvostrano)').font = n; ws.cell(row=r + 4, column=2, value=f'=IF(B{r+3}="","",2*(1-NORMSDIST(ABS(B{r+3}))))').number_format = '0.000'
ws.cell(row=r + 5, column=1, value='Tumačenje: p > 0,05 → nema dokaza da se preciznost mijenjala kroz vrijeme; trendovi i udjeli tada nisu pristrani zbog šuma upita.').font = small
ws.cell(row=r + 7, column=1, value='Napomena: formule se odnose na stupac L (KONAČNA OCJENA) oba lista; prazne ćelije se ne broje.').font = small
ws.column_dimensions['A'].width = 62; ws.column_dimensions['B'].width = 14

# ---------------- strict query sensitivity ----------------
ws = wb.create_sheet('Strogi_upit')
ws.cell(row=1, column=1, value='Osjetljivost na strogi upit: D AND E NOT (crop OR crops OR yield OR agriculture OR agricultural OR wheat OR maize OR rice OR soybean OR vineyard OR phenotyping OR weed OR weeds); OpenAlex, dohvat 27. 8. 2026.').font = Font(name=FONT, size=10, bold=True)
core = {2005:28,2006:37,2007:43,2008:44,2009:55,2010:55,2011:75,2012:105,2013:157,2014:185,2015:330,2016:507,2017:642,2018:906,2019:1248,2020:1596,2021:1959,2022:2106,2023:2369,2024:3150,2025:4747}
strict = {2005:24,2006:35,2007:38,2008:38,2009:48,2010:48,2011:65,2012:84,2013:124,2014:132,2015:266,2016:394,2017:502,2018:659,2019:926,2020:1155,2021:1445,2022:1550,2023:1696,2024:2188,2025:3309}
for c, h in enumerate(['Godina', 'core_drone_env', 'strict_noagri', 'omjer strict/core'], 1):
    cell = ws.cell(row=3, column=c, value=h); cell.font = b; cell.fill = hdr
for i, y in enumerate(range(2005, 2026), start=4):
    ws.cell(row=i, column=1, value=y).font = n
    ws.cell(row=i, column=2, value=core[y]).font = blue; ws.cell(row=i, column=3, value=strict[y]).font = blue
    ws.cell(row=i, column=4, value=f'=C{i}/B{i}').number_format = '0.00'
ws.cell(row=26, column=1, value='CAGR 2005–2025').font = b
ws.cell(row=26, column=2, value='=(B24/B4)^(1/20)-1').number_format = '0.0%'; ws.cell(row=26, column=3, value='=(C24/C4)^(1/20)-1').number_format = '0.0%'
ws.cell(row=27, column=1, value='CAGR 2015–2025').font = b
ws.cell(row=27, column=2, value='=(B24/B14)^(1/10)-1').number_format = '0.0%'; ws.cell(row=27, column=3, value='=(C24/C14)^(1/10)-1').number_format = '0.0%'
ws.cell(row=29, column=1, value='Tumačenje: omjer pada s ~0,86 (2005.) na ~0,70 (2025.) — udio poljoprivrednih radova u upitu raste; strogi upit raste ~1,4 postotna boda sporije godišnje. Trend i zaključak o zasićenju udjela ostaju isti (provjeriti u MATLAB-u s opt.numerator = \'strict_noagri\').').font = small
ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 16; ws.column_dimensions['C'].width = 16; ws.column_dimensions['D'].width = 18

wb.save('Provjera_preciznosti_upita.xlsx'); print('saved')
