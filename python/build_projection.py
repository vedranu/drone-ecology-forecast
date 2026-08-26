"""Assemble the 10-year projection deliverables from ../projection/*/ outputs:
   - ../projection/Projekcija_2026_2035.xlsx
   - ../projection/fig_projection_subdomains.png
Run after drone_env_projection_10y.m (or the Octave equivalent)."""
import os, pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib; matplotlib.use('Agg')
import matplotlib.pyplot as plt

P = '../projection/'; D = '../data/'
annual = pd.read_csv(D + 'openalex_drone_env_annual.csv')
subs = ['sub_fauna', 'sub_forest', 'sub_water', 'sub_air', 'sub_wildfire', 'sub_conservation']
sub_lab = {'sub_fauna': 'Fauna (brojanje/popis)', 'sub_forest': 'Šume / zdravlje šuma', 'sub_water': 'Vode / bentos',
           'sub_air': 'Zrak / emisije plinova', 'sub_wildfire': 'Požari', 'sub_conservation': 'Zaštićena područja'}
mlab = {'exponential': 'Eksponencijalni', 'logistic': 'Logistički', 'gompertz': 'Gompertz', 'bass': 'Bass', 'qaicc_average': 'QAICc-prosjek'}

def load(name):
    fc = pd.read_csv(P + name + '/forecast_count.csv'); fs = pd.read_csv(P + name + '/forecast_share.csv')
    mc = pd.read_csv(P + name + '/models_count.csv'); ms = pd.read_csv(P + name + '/models_share.csv')
    return fc, fs, mc, ms

FONT = 'Arial'
hdr_fill = PatternFill('solid', fgColor='D9E2F3'); avg_fill = PatternFill('solid', fgColor='FFF2CC')
b_font = Font(name=FONT, size=10, bold=True); n_font = Font(name=FONT, size=10); in_font = Font(name=FONT, size=10, color='0000FF')
thin = Side(style='thin', color='BFBFBF')
wb = Workbook()

def header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c); cell.font = b_font; cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center'); cell.border = Border(bottom=thin)

def widths(ws, w):
    for i, x in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = x

def wide(fc, fmt_scale=1):
    """pivot: rows = year, columns = model point / avg q05 q50 q95"""
    years = sorted(fc.year.unique())
    out = pd.DataFrame(index=years)
    for m in ['exponential', 'logistic', 'gompertz', 'bass']:
        sub = fc[fc.model == m]
        if len(sub): out[m] = sub.set_index('year')['point'] * fmt_scale
    a = fc[fc.model == 'qaicc_average'].set_index('year')
    out['avg_q50'] = a['q50'] * fmt_scale; out['avg_q05'] = a['q05'] * fmt_scale; out['avg_q95'] = a['q95'] * fmt_scale
    return out

# ---------------- core sheets ----------------
fcC, fsC, mcC, msC = load('core')
wC = wide(fcC); wS = wide(fsC)          # share CSV already in per-1000 units
obs = annual[(annual.year >= 2015) & (annual.year <= 2025)]

def sheet_projection(title, w, obs_series, fmt, unit, note, models_present):
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=f'{title} — {unit}').font = Font(name=FONT, size=11, bold=True)
    ws.cell(row=2, column=1, value=note).font = Font(name=FONT, size=9, italic=True)
    cols = ['Godina', 'Tip'] + [mlab[m] for m in models_present] + ['Prosjek modela (medijan)', 'Donja granica 90 %', 'Gornja granica 90 %']
    r = 4
    for c, h in enumerate(cols, 1): ws.cell(row=r, column=c, value=h)
    header(ws, r, len(cols)); r += 1
    for y, v in obs_series.items():
        ws.cell(row=r, column=1, value=int(y)).font = n_font; ws.cell(row=r, column=2, value='opaženo').font = n_font
        c = ws.cell(row=r, column=3 + len(models_present), value=float(v)); c.font = in_font; c.number_format = fmt; r += 1
    for y, row in w.iterrows():
        ws.cell(row=r, column=1, value=int(y)).font = n_font; ws.cell(row=r, column=2, value='projekcija').font = n_font
        for i, m in enumerate(models_present):
            c = ws.cell(row=r, column=3 + i, value=float(row[m])); c.font = n_font; c.number_format = fmt
        for j, k in enumerate(['avg_q50', 'avg_q05', 'avg_q95']):
            c = ws.cell(row=r, column=3 + len(models_present) + j, value=float(row[k])); c.font = b_font if j == 0 else n_font
            c.number_format = fmt; c.fill = avg_fill
        r += 1
    ws.freeze_panes = 'C5'; widths(ws, [9, 11] + [15] * len(models_present) + [22, 18, 18])
    return ws

sheet_projection('Core_broj_radova', wC, obs.set_index('year')['core_drone_env'], '#,##0', 'broj radova godišnje (OpenAlex, upit D AND E)',
                 'Točkaste procjene četiriju modela i QAICc-ponderirani prosjek s 90 % negativno-binomnim bootstrap intervalom (1000 replika). Opaženo 2015–2025 radi konteksta.',
                 ['exponential', 'logistic', 'gompertz', 'bass'])
obs_share = (obs.set_index('year')['core_drone_env'] / obs.set_index('year')['env_only'] * 1000)
sheet_projection('Core_udio', wS, obs_share, '0.00', 'radova na 1 000 radova o okolišu',
                 'Modeli stope s nazivnikom env_only kao offsetom (Bass se ne koristi za udio). Prosjek i 90 % interval iz 1000 replika.',
                 ['exponential', 'logistic', 'gompertz'])

# ---------------- subdomains ----------------
ws = wb.create_sheet('Poddomene')
ws.cell(row=1, column=1, value='Projekcija poddomena 2026–2035 — broj radova godišnje (QAICc-prosjek modela, medijan i 90 % interval, 200 replika)').font = Font(name=FONT, size=11, bold=True)
ws.cell(row=2, column=1, value='Poddomenski upiti se preklapaju (D AND poddomenski pojmovi) i nisu podskup core upita; vidi codebook.md.').font = Font(name=FONT, size=9, italic=True)
cols = ['Godina', 'Tip'] + [f'{sub_lab[s]}{sfx}' for s in subs for sfx in ('', ' q05', ' q95')]
r = 4
for c, h in enumerate(cols, 1): ws.cell(row=r, column=c, value=h)
header(ws, r, len(cols)); r += 1
subw = {}
weights = {}
for s in subs:
    fc, fs, mc, ms = load(s); subw[s] = wide(fc); weights[s] = mc
for y in range(2015, 2036):
    ws.cell(row=r, column=1, value=y).font = n_font
    ws.cell(row=r, column=2, value='opaženo' if y <= 2025 else 'projekcija').font = n_font
    for i, s in enumerate(subs):
        if y <= 2025:
            v = int(annual.loc[annual.year == y, s].iloc[0])
            c = ws.cell(row=r, column=3 + 3 * i, value=v); c.font = in_font; c.number_format = '#,##0'
        else:
            row = subw[s].loc[y]
            for j, k in enumerate(['avg_q50', 'avg_q05', 'avg_q95']):
                c = ws.cell(row=r, column=3 + 3 * i + j, value=float(row[k])); c.number_format = '#,##0'
                c.font = b_font if j == 0 else n_font; c.fill = avg_fill if j == 0 else PatternFill()
    r += 1
ws.freeze_panes = 'C5'; widths(ws, [9, 11] + [14, 10, 10] * len(subs))

# ---------------- model weights ----------------
ws = wb.create_sheet('Modeli_tezine')
ws.cell(row=1, column=1, value='Modeli, QAICc-težine i backtest (MAPE %) po seriji').font = Font(name=FONT, size=11, bold=True)
cols = ['Serija', 'Model', 'k', 'Devijanca', 'QAICc', 'Težina', 'Backtest MAPE %', 'Parametri']
r = 3
for c, h in enumerate(cols, 1): ws.cell(row=r, column=c, value=h)
header(ws, r, len(cols)); r += 1
for name, mc in [('core_drone_env (broj)', mcC), ('core_drone_env (udio)', msC)] + [(s, weights[s]) for s in subs]:
    for row in mc.itertuples(index=False):
        vals = [name, mlab.get(row.model, row.model), int(row.k), float(row.deviance), float(row.QAICc), float(row.weight), float(row.backtest_MAPE), row.parameters]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.font = n_font
            if c in (4, 5): cell.number_format = '0.0'
            if c == 6: cell.number_format = '0.000'
            if c == 7: cell.number_format = '0.0'
        r += 1
widths(ws, [24, 16, 5, 11, 10, 9, 15, 50])

# ---------------- notes ----------------
ws = wb.create_sheet('Napomene')
notes = [
 'Metoda: četiri modela rasta (eksponencijalni, logistički, Gompertz, Bass) prilagođena Poissonovom devijancom godišnjim brojevima 2005–2025; izbor i ponderiranje modela QAICc-om (predisperzija ĉ); nesigurnost parametarskim negativno-binomnim bootstrapom; rolling-origin backtest (ishodišta 2018, 2020, 2022).',
 'Skripta: matlab/drone_env_projection_10y.m (poziva drone_env_forecast.m s horizontima 1–10). Ovi rezultati su iz GNU Octave 8.4; MATLAB daje iste točkaste procjene, neznatno drukčije intervale.',
 'Točkasta procjena "Prosjek modela" je medijan bootstrap-distribucije QAICc-ponderiranog prosjeka; "Donja/Gornja granica" su 5. i 95. percentil.',
 'Interpretacija: horizont do 3 godine je robustan (modeli se slažu), horizont 7–10 godina ovisi o tome je li serija prije ili na točki infleksije — što podaci do 2025. ne mogu razlučiti. Udio u literaturi o okolišu ima jasnu logističku saturaciju (~13 na 1 000).',
 'Podaci: OpenAlex API, dohvat 26. 8. 2026.; 2026 je nepotpuna i nije korištena u fitu. Skok u 2025. dijelom je artefakt indeksiranja (vidi codebook.md, analiza osjetljivosti u results/console_output).',
 'Žuto = QAICc-prosjek (preporučena vrijednost za citiranje), plavo = opaženi podaci.',
]
for i, t in enumerate(notes, 1):
    c = ws.cell(row=i, column=1, value=t); c.font = n_font; c.alignment = Alignment(wrap_text=True, vertical='top')
ws.column_dimensions['A'].width = 150
del wb['Sheet']
wb.save(P + 'Projekcija_2026_2035.xlsx')

# ---------------- figure: subdomain projections ----------------
fig, axes = plt.subplots(2, 3, figsize=(7.2, 4.6), sharex=True)
for ax, s in zip(axes.flat, subs):
    o = annual[(annual.year >= 2010) & (annual.year <= 2025)]
    w = subw[s]
    ax.fill_between(w.index, w.avg_q05, w.avg_q95, color='0.8', lw=0, label='90 % interval')
    ax.plot(o.year, o[s], 'ko', ms=2.5, label='opaženo')
    ax.plot(w.index, w.avg_q50, 'k-', lw=1.2, label='prosjek modela')
    ax.set_title(sub_lab[s], fontsize=8); ax.set_yscale('log'); ax.tick_params(labelsize=7); ax.grid(True, which='major', lw=0.3)
    ax.set_xlim(2010, 2035)
axes[0, 0].legend(fontsize=6, loc='upper left')
for ax in axes[1]: ax.set_xlabel('Godina', fontsize=7)
for ax in axes[:, 0]: ax.set_ylabel('Radova godišnje', fontsize=7)
fig.tight_layout(); fig.savefig(P + 'fig_projection_subdomains.png', dpi=300); plt.close(fig)

# ---------------- figure: core year-by-year ----------------
fig, ax = plt.subplots(figsize=(4.5, 3.4))
o = annual[(annual.year >= 2010) & (annual.year <= 2025)]
ax.fill_between(wC.index, wC.avg_q05, wC.avg_q95, color='0.8', lw=0, label='90 % interval (prosjek)')
for m, st in zip(['exponential', 'logistic', 'gompertz', 'bass'], ['-', '--', '-.', ':']):
    ax.plot(wC.index, wC[m], 'k' + st, lw=0.8, label=mlab[m])
ax.plot(o.year, o.core_drone_env, 'ko', ms=3, label='opaženo')
ax.plot(wC.index, wC.avg_q50, 'ks', mfc='w', ms=4, label='prosjek modela')
ax.set_yscale('log'); ax.set_xlabel('Godina', fontsize=8); ax.set_ylabel('Radova godišnje (OpenAlex)', fontsize=8)
ax.tick_params(labelsize=7); ax.grid(True, lw=0.3); ax.legend(fontsize=6, loc='upper left')
fig.tight_layout(); fig.savefig(P + 'fig_projection_core.png', dpi=300); plt.close(fig)
print(wC.round(0)); print(wS.round(2))
print('done')
